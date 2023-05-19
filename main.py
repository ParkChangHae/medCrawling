import csv
import json
import random
import re
import time
import traceback
from collections import defaultdict
from datetime import datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common import StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from seleniumwire import webdriver
# A package to have a chromedriver always up-to-date.
from webdriver_manager.chrome import ChromeDriverManager
from fake_useragent import UserAgent

FINAL_SAVE_XLSX_FILE_NAME = './경쟁사_QnA게시판_현황.xlsx'
DEFAULT_THRESHOLD_DATE = datetime(2023, 5, 5) #이 날짜를 포함한 날 까지 크롤링.
TEACHER_URL_INFO = './강사 정보.csv'


def load_from_xslx(teacher_info_dict):

    try:
        wb = load_workbook(FINAL_SAVE_XLSX_FILE_NAME)

        qna_Sheet = wb['QnA게시판 현황']
        metadata_Sheet = wb['metadata']

        #qna시트로부터 기존에 질문 cnt load
        i = 2
        while(True):
            site_name = qna_Sheet.cell(row = 1, column=i).value
            subject = qna_Sheet.cell(row = 2, column=i).value
            teacher_name = qna_Sheet.cell(row = 3, column=i).value

            if(site_name == None or subject == None or teacher_name == None):
                break

            j = 4
            while(True):
                date = qna_Sheet.cell(row = j, column=1).value
                date_qna_cnt = qna_Sheet.cell(row = j, column=i).value

                if(date_qna_cnt == None):
                    date_qna_cnt = 0

                if(date == None):
                    break

                teacher_info_dict[site_name][subject][teacher_name]['qna_cnt'][date]['총질문수'] = date_qna_cnt

                j+=1


            i+=1

        #start_date, end_date load
        i = 1
        while(True):
            site_name = metadata_Sheet.cell(row = 1, column=i).value
            subject = metadata_Sheet.cell(row = 2, column=i).value
            teacher_name = metadata_Sheet.cell(row = 3, column=i).value
            start_date = metadata_Sheet.cell(row = 4, column=i).value
            end_date = metadata_Sheet.cell(row = 5, column=i).value

            if(site_name == None or subject == None or teacher_name == None):
                break

            if(start_date == None):
                start_date = DEFAULT_THRESHOLD_DATE
            if(end_date == None):
                end_date = DEFAULT_THRESHOLD_DATE

            teacher_info_dict[site_name][subject][teacher_name]['start_date'] = start_date
            teacher_info_dict[site_name][subject][teacher_name]['end_date'] = end_date

            #start_date, end_date는 매번 새롭게 크롤링 진행함.
            teacher_info_dict[site_name][subject][teacher_name]['qna_cnt'][start_date]['총질문수'] = 0
            teacher_info_dict[site_name][subject][teacher_name]['qna_cnt'][end_date]['총질문수'] = 0

            i+=1

        latest_date = qna_Sheet['A4'].value
        today_date = datetime.now()

        time_gap = (today_date - latest_date).days

        if (time_gap > 0):
            qna_Sheet.insert_rows(4, time_gap)  # 새롭게 생성해야하는 만큼 Row insert!
        elif (time_gap == 0):
            pass
        else:
            pass
            # ERROR!!! 발생 불가능 케이스

        wb.save(FINAL_SAVE_XLSX_FILE_NAME)

    except Exception:
        pass
        # 파일 존재하지 않을 때

def save_to_xslx(teacher_info_dict):
    try:
        wb = load_workbook(FINAL_SAVE_XLSX_FILE_NAME)
        qna_Sheet = wb['QnA게시판 현황']
        metadata_sheet = wb['metadata']

    except Exception:
        # 파일이 없는 경우
        wb = Workbook()
        qna_Sheet = wb.active
        qna_Sheet.title = "QnA게시판 현황"
        metadata_sheet = wb.create_sheet('metadata')

    # Set the start and end dates
    now = datetime.now()
    start_date = datetime(now.year, now.month, now.day)

    # Start from the cell A4
    row = 4
    col = 1

    current_date = start_date
    while current_date >= DEFAULT_THRESHOLD_DATE:
        qna_Sheet.cell(row=row, column=col).value = current_date
        qna_Sheet.cell(row=row, column=col).number_format = 'YYYY/MM/DD'
        current_date -= timedelta(days=1)
        row += 1  # move to the next row

    j = 2
    for site_name in teacher_info_dict.keys():
        for subject in teacher_info_dict[site_name].keys():
            for teacher in teacher_info_dict[site_name][subject].keys():
                current_teacher_dict = teacher_info_dict[site_name][subject][teacher]

                # 가로줄  #세로줄
                qna_Sheet.cell(row=1, column=j).value = site_name
                qna_Sheet.cell(row=2, column=j).value = subject
                qna_Sheet.cell(row=3, column=j).value = teacher

                metadata_sheet.cell(row=1, column=j-1).value = site_name
                metadata_sheet.cell(row=2, column=j-1).value = subject
                metadata_sheet.cell(row=3, column=j-1).value = teacher

                if(len(current_teacher_dict['qna_cnt']) > 0):
                    metadata_sheet.cell(row=4, column=j-1).value = max(current_teacher_dict['qna_cnt'])
                    metadata_sheet.cell(row=5, column=j-1).value = min(current_teacher_dict['qna_cnt'])
                else :
                    metadata_sheet.cell(row=4, column=j-1).value = DEFAULT_THRESHOLD_DATE
                    metadata_sheet.cell(row=5, column=j-1).value = DEFAULT_THRESHOLD_DATE

                metadata_sheet.cell(row=4, column=j - 1).number_format = 'YYYY/MM/DD'
                metadata_sheet.cell(row=5, column=j - 1).number_format = 'YYYY/MM/DD'

                for date in current_teacher_dict['qna_cnt'].keys():

                    offset = (start_date - date).days + 4

                    qna_Sheet.cell(row=offset, column=j).value = \
                        teacher_info_dict[site_name][subject][teacher]['qna_cnt'][date]['총질문수']

                j += 1

    #sheet_metadata.sheet_state = 'hidden'

    # Save the workbook to a file
    wb.save(FINAL_SAVE_XLSX_FILE_NAME)


def week_of_month(dt):
    first_day_of_month = dt.replace(day=1)
    day_of_month = dt.day
    adjusted_dom = day_of_month + first_day_of_month.weekday()

    return (adjusted_dom - 1) // 7 + 1


def load_teacher_info(file_name='강사 정보.csv'):
    teacher_info_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))

    with open(file_name, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        for row in reader:
            teacher_info_dict[row[0]][row[1]][row[2]] = {'url': row[3],
                                                         'qna_cnt': defaultdict(lambda: defaultdict(int)), 'start_date' : DEFAULT_THRESHOLD_DATE, 'end_date' : DEFAULT_THRESHOLD_DATE, 'threshold_date' : DEFAULT_THRESHOLD_DATE }


    return teacher_info_dict


def is_contains_date(text):
    # 정규 표현식 패턴 정의
    date_pattern = re.compile(r'\d{4}/\d{2}/\d{2}')
    date_pattern2 = re.compile(r'\d{4}.\d{2}.\d{2}')

    # 텍스트에서 날짜를 찾음
    match = date_pattern.search(text)
    match2 = date_pattern2.search(text)

    # 날짜가 발견되면 True 반환, 그렇지 않으면 False 반환
    return (match is not None) or (match2 is not None)


def is_match_date(text):
    # 정규 표현식 패턴 정의
    date_pattern = re.compile(r'\d{4}/\d{2}/\d{2}')
    date_pattern2 = re.compile(r'\d{4}.\d{2}.\d{2}')

    # 텍스트에서 날짜를 찾음
    match = date_pattern.match(text)
    match2 = date_pattern2.match(text)

    # 날짜가 발견되면 True 반환, 그렇지 않으면 False 반환
    return (match is not None) or (match2 is not None)


def contains_keyword(text, keyword):
    pattern = r'\b' + keyword + r'\b'
    if re.search(pattern, text):
        return True
    return False


def save_to_dict(current_teacher_dict, date, status="답변완료"):
    status = '총질문수'
    current_teacher_dict['qna_cnt'][date][status] += 1
    #current_teacher_dict['end_date'] = date


class NextTeacherRaise(Exception):
    pass

def presume_start_num(site_name, qna_sum):

    if(site_name == "메가스터디"):
        iter = int(qna_sum / 500) - 1

    elif(site_name == "이투스"):
        iter = int(qna_sum / 10) - 10

    elif(site_name == "대성마이맥"):
        iter = int(qna_sum * 2 / 5000) - 1

    if(iter <= 0):
        iter = 1

    return iter


def get_data(browser, teacher_info_dict, site_name, subject, teacher_name, iter = 1):

    MAX_RETRIES = 10

    current_teacher_dict = teacher_info_dict[site_name][subject][teacher_name]
    original_qna_url = current_teacher_dict['url']

    while(True):
        try:
            print("i : " + str(iter))
            retries = 0
            while retries < MAX_RETRIES:
                try:

                    if (site_name == "메가스터디"):

                        bring_amount = 500

                        target_url = original_qna_url + "startRowNum=" + str(
                            bring_amount * (iter - 1)) + "&itemPerPage=" + str(bring_amount)

                        start_time = time.time()  # 코드 실행 전의 시간

                        browser.get(target_url)
                        browser.implicitly_wait(10)

                        end_time = time.time()  # 코드 실행 후의 시간
                        execution_time = end_time - start_time  # 실행 시간 계산
                        print(f"데이터 크롤링 시간 : {execution_time}초")

                        elements = WebDriverWait(browser, 20).until(
                            EC.presence_of_element_located((By.TAG_NAME, 'body')))

                        try:
                            match = re.search(r'\"aData\":(.*?), \"bData\"', elements.text, re.DOTALL)
                            raw_json_data = match.group(1)
                            fixed_elements_text = re.sub(r'"BT_TITLE":.*?"QNAFLG":".*?",', '',
                                                         raw_json_data)
                            json_data = json.loads(fixed_elements_text)
                        except Exception:
                            raise StaleElementReferenceException

                        for data_element in json_data:

                            if (int(data_element['CNT']) > 0):
                                status = "답변완료"
                            else:
                                status = "답변대기"

                            date = datetime.strptime(data_element['REG_DT'], "%Y-%m-%d")

                            if (date < current_teacher_dict['threshold_date']):
                                raise NextTeacherRaise
                                # 다음 선생님으로 넘어가야함

                            if (date >= current_teacher_dict['start_date'] or date <= current_teacher_dict['end_date']):
                                save_to_dict(current_teacher_dict, date, status)

                        if (len(json_data) <= 0):
                            raise NextTeacherRaise

                    elif (site_name == "이투스"):

                        target_url = original_qna_url + str(iter)

                        start_time = time.time()  # 코드 실행 전의 시간

                        browser.get(target_url)
                        browser.implicitly_wait(10)

                        end_time = time.time()  # 코드 실행 후의 시간
                        execution_time = end_time - start_time  # 실행 시간 계산
                        print(f"데이터 크롤링 시간 : {execution_time}초")

                        elements = WebDriverWait(browser, 20).until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, 'ul.board_list')))

                        origin_text_list = str(elements.text).split("\n")

                        for j in range(len(origin_text_list) - 1, -1, -1):

                            if (origin_text_list[j] == '대기' or origin_text_list[j] == '완료'):
                                pass

                            elif (contains_keyword(origin_text_list[j], '공지') or not is_contains_date(
                                    origin_text_list[j]) or is_match_date(origin_text_list[j])):
                                origin_text_list.pop(j)

                        for j in range(0, len(origin_text_list), 2):

                            if (origin_text_list[j] == '완료'):
                                status = "답변완료"
                            else:
                                status = "답변대기"

                            date = origin_text_list[j + 1].split()[-1:][0].replace('.', '-')

                            date = datetime.strptime(date, "%Y-%m-%d")

                            if (len(origin_text_list) <= 0 or date < current_teacher_dict[
                                'threshold_date']):
                                raise NextTeacherRaise
                                # 다음 선생님으로 넘어가야함

                            if (date >= current_teacher_dict['start_date'] or date <= current_teacher_dict['end_date']):
                                save_to_dict(current_teacher_dict, date, status)

                    elif (site_name == "대성마이맥"):

                        ua = UserAgent()
                        userAgnet = ua.random

                        USERNAME = "fpdlwj4869"
                        PASSWORD = "Wlrnrhkgkr1!"

                        entry = ('http://customer-%s:%s@dc.kr-pr.oxylabs.io:14000' %
                                 (USERNAME, PASSWORD))

                        proxy_dict = {
                            "http": entry,
                            "https": entry
                        }

                        headers = {
                            "Origin": "https://m.mimacstudy.com",
                            "Referer": "https://m.mimacstudy.com/mobile/tcher/studyQna.ds?coType=M&tcd=" +
                                       current_teacher_dict['url'],
                            "User-Agent": userAgnet
                        }

                        params = {
                            "pid": "",
                            "tcd": current_teacher_dict['url'],
                            "currPage": str(iter),
                            "pagePerCnt": "5000",
                            "myQna": "N",
                            "srchWordTxt": "",
                            "isScrtN": "N"
                        }

                        ipaddress = requests.get(url="https://ip.oxylabs.io/", proxies=proxy_dict)
                        print("IP : " + ipaddress.text + " User Agent : " + userAgnet)

                        start_time = time.time()  # 코드 실행 전의 시간

                        while (True):
                            res = requests.get(
                                url="https://m.mimacstudy.com/mobile/tcher/getStudyQnaListByAjax.ds",
                                headers=headers, params=params, proxies=proxy_dict)
                            json_data = json.loads(res.text)

                            if (json_data['code'] == 'success'):
                                json_data = json_data['data']
                                break

                        end_time = time.time()  # 코드 실행 후의 시간
                        execution_time = end_time - start_time  # 실행 시간 계산
                        print(f"데이터 크롤링 시간 : {execution_time}초")

                        if (len(json_data) <= 0):
                            raise NextTeacherRaise

                        for data_element in json_data:

                            if (data_element['qnaType'] == 'Q'):
                                status = "답변완료"
                            else:
                                status = "답변완료"
                                continue

                            date = data_element['regDate'].replace('/', '-')
                            date = datetime.strptime(date, "%Y-%m-%d")

                            if (date < current_teacher_dict['threshold_date']):
                                raise NextTeacherRaise
                                # 다음 선생님으로 넘어가야함

                            if (date >= current_teacher_dict['start_date'] or date <= current_teacher_dict['end_date']):
                                save_to_dict(current_teacher_dict, date, status)

                        print("xx")

                    break

                except StaleElementReferenceException:
                    retries += 1
                    print("Exception : " + str(retries))
                    continue

            if (iter % 50 == 0):

                save_to_xslx(teacher_info_dict)

                print(iter)
                for date, status_counts in current_teacher_dict['qna_cnt'].items():
                    print(
                        f"{site_name}_{subject}_{teacher_name} | {date}: 총 질문 수 {str(int(status_counts['총질문수']))}개")

            # save_to_xslx(teacher_info_dict)

        except NextTeacherRaise:

            save_to_xslx(teacher_info_dict)

            for date, status_counts in current_teacher_dict['qna_cnt'].items():
                print(
                    f"{site_name}_{subject}_{teacher_name} | {date}: 총 질문 수 {str(int(status_counts['총질문수']))}개")

            break

        except Exception as e:

            current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            with open(f"{current_time}_Exception.txt", "w") as file:
                file.write(str(e))
                file.write(traceback.format_exc())

            print(e)
        finally:
            iter+=1


def crawling_qna(teacher_info_dict):

    chrome_options = Options()
    chrome_options.add_argument('--headless')  # 옵션: 창이 나타나지 않게 함
    chrome_options.add_argument("--enable-javascript")
    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    load_from_xslx(teacher_info_dict)

    for site_name in teacher_info_dict.keys():
        for subject in teacher_info_dict[site_name].keys():
            for teacher_name in teacher_info_dict[site_name][subject].keys():
                current_teacher_dict = teacher_info_dict[site_name][subject][teacher_name]

                #Today ~ start_date 까지 탐색

                current_teacher_dict['threshold_date'] = current_teacher_dict['start_date']
                get_data(browser, teacher_info_dict, site_name, subject, teacher_name, iter=1)

                if( current_teacher_dict['end_date'] >= DEFAULT_THRESHOLD_DATE ):

                    total_questions = sum(inner_dict['총질문수'] for inner_dict in current_teacher_dict['qna_cnt'].values())
                    iter = presume_start_num(site_name, total_questions )

                    current_teacher_dict['threshold_date'] = DEFAULT_THRESHOLD_DATE
                    get_data(browser, teacher_info_dict, site_name, subject, teacher_name, iter)

    save_to_xslx(teacher_info_dict)

    # 웹 드라이버 종료
    browser.quit()


def main():
    teacher_info_dict = load_teacher_info('강사 정보_테스트용.csv')
    crawling_qna(teacher_info_dict)
    return 0


if __name__ == '__main__':
    main()
